"""
Record an experiment run (train or eval) into a shared Excel tracker.

Usage
-----
    python scripts/record_run.py \
        --xlsx experiments.xlsx \
        --run-name BH-000425-08-05-VDA-V3-KITTI-lr1e-5 \
        --status done \
        --job-id 12200409 \
        --encoder vitl \
        --lr 1e-5 \
        --batch-size 4 \
        --epochs 25 \
        --log-dir logs/depth_anything_v3/kitti_eigen_vitl \
        --metrics-json logs/.../test_metrics.json \
        --notes "LR sweep, frozen 5 epochs then unfreeze"

Columns (sheet: 'runs'):
    run_name | project_code | stage | status | job_id | submitted_at | finished_at
    encoder | freeze_backbone | freeze_until | lr | backbone_lr_factor | batch_size
    max_epochs | precision | input_hw | dataset_split | ckpt
    abs_rel | sq_rel | rmse | rmse_log | d1 | d2 | d3 | log_dir | notes

The file is created on first call; subsequent calls append a new row (or update an
existing row matching --run-name + --stage if --update is passed).
Uses filelock to make concurrent bsub-tail calls safe.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
from pathlib import Path
from typing import Any, Dict, Optional

from filelock import FileLock
from openpyxl import Workbook, load_workbook


COLUMNS = [
    "run_name", "project_code", "stage", "status", "job_id",
    "submitted_at", "finished_at",
    "encoder", "freeze_backbone", "freeze_until",
    "lr", "backbone_lr_factor", "batch_size", "max_epochs",
    "precision", "input_hw", "dataset_split", "ckpt",
    "abs_rel", "sq_rel", "rmse", "rmse_log", "d1", "d2", "d3",
    "log_dir", "notes",
]

PROJECT_CODE = "BH-000425-08-05"


def _now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _load_metrics(metrics_json: Optional[str]) -> Dict[str, float]:
    if not metrics_json:
        return {}
    p = Path(metrics_json)
    if not p.is_file():
        print(f"[record] WARN: metrics file not found: {p}")
        return {}
    with p.open("r") as f:
        data = json.load(f)
    # Accept flat {abs_rel:...} or nested {test:{abs_rel:...}} / with 'test/abs_rel' keys
    out: Dict[str, float] = {}
    mapping_keys = ["abs_rel", "sq_rel", "rmse", "rmse_log", "d1", "d2", "d3"]
    for k in mapping_keys:
        for candidate in (k, f"test/{k}", f"val/{k}"):
            if candidate in data:
                try:
                    out[k] = float(data[candidate])
                except Exception:
                    pass
                break
    return out


def _ensure_workbook(xlsx_path: Path) -> Workbook:
    if xlsx_path.exists():
        return load_workbook(xlsx_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "runs"
    ws.append(COLUMNS)
    # Simple header styling (bold)
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    # Freeze header row
    ws.freeze_panes = "A2"
    return wb


def _find_row(ws, run_name: str, stage: str) -> Optional[int]:
    """Return 1-based row index for a matching (run_name, stage), or None."""
    for row_idx in range(2, ws.max_row + 1):
        rn = ws.cell(row=row_idx, column=1).value
        st = ws.cell(row=row_idx, column=3).value
        if rn == run_name and st == stage:
            return row_idx
    return None


def _write_row(ws, row_idx: int, record: Dict[str, Any]) -> None:
    for col_idx, key in enumerate(COLUMNS, start=1):
        val = record.get(key, None)
        if val is None or val == "":
            continue
        ws.cell(row=row_idx, column=col_idx, value=val)


def build_record(args: argparse.Namespace) -> Dict[str, Any]:
    metrics = _load_metrics(args.metrics_json)
    rec: Dict[str, Any] = {
        "run_name": args.run_name,
        "project_code": args.project_code or PROJECT_CODE,
        "stage": args.stage,
        "status": args.status,
        "job_id": args.job_id,
        "submitted_at": args.submitted_at,
        "finished_at": args.finished_at or (_now_str() if args.status in ("done", "failed") else ""),
        "encoder": args.encoder,
        "freeze_backbone": args.freeze_backbone,
        "freeze_until": args.freeze_until,
        "lr": args.lr,
        "backbone_lr_factor": args.backbone_lr_factor,
        "batch_size": args.batch_size,
        "max_epochs": args.max_epochs,
        "precision": args.precision,
        "input_hw": args.input_hw,
        "dataset_split": args.dataset_split,
        "ckpt": args.ckpt,
        "log_dir": args.log_dir,
        "notes": args.notes,
    }
    rec.update({k: metrics.get(k) for k in ("abs_rel", "sq_rel", "rmse", "rmse_log", "d1", "d2", "d3")})
    return rec


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=str, default="experiments.xlsx")
    p.add_argument("--run-name", type=str, required=True)
    p.add_argument("--project-code", type=str, default=PROJECT_CODE)
    p.add_argument("--stage", type=str, default="train",
                   choices=["train", "eval", "baseline"])
    p.add_argument("--status", type=str, default="submitted",
                   choices=["submitted", "running", "done", "failed"])
    p.add_argument("--job-id", type=str, default="")
    p.add_argument("--submitted-at", type=str, default="")
    p.add_argument("--finished-at", type=str, default="")

    p.add_argument("--encoder", type=str, default="")
    p.add_argument("--freeze-backbone", type=str, default="")
    p.add_argument("--freeze-until", type=str, default="")
    p.add_argument("--lr", type=str, default="")
    p.add_argument("--backbone-lr-factor", type=str, default="")
    p.add_argument("--batch-size", type=str, default="")
    p.add_argument("--max-epochs", type=str, default="")
    p.add_argument("--precision", type=str, default="")
    p.add_argument("--input-hw", type=str, default="")
    p.add_argument("--dataset-split", type=str, default="")
    p.add_argument("--ckpt", type=str, default="")

    p.add_argument("--metrics-json", type=str, default="")
    p.add_argument("--log-dir", type=str, default="")
    p.add_argument("--notes", type=str, default="")

    p.add_argument("--update", action="store_true",
                   help="If a row with same run_name+stage exists, update it instead of appending.")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    lock = FileLock(str(xlsx_path) + ".lock", timeout=60)
    record = build_record(args)
    with lock:
        wb = _ensure_workbook(xlsx_path)
        ws = wb["runs"] if "runs" in wb.sheetnames else wb.active

        row_idx: Optional[int] = None
        if args.update:
            row_idx = _find_row(ws, args.run_name, args.stage)
        if row_idx is None:
            row_idx = ws.max_row + 1
        _write_row(ws, row_idx, record)
        wb.save(xlsx_path)

    print(f"[record] wrote row {row_idx} → {xlsx_path}")
    print(f"[record] run_name={args.run_name} stage={args.stage} status={args.status}")
    for k in ("abs_rel", "sq_rel", "rmse", "rmse_log", "d1", "d2", "d3"):
        v = record.get(k)
        if v is not None:
            print(f"  {k:<10s} = {v}")


if __name__ == "__main__":
    main()
